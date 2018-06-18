import csv
import operator

from openpyxl import Workbook
from openpyxl.styles import Alignment
from openpyxl.styles.borders import Border, Side
from openpyxl.styles import PatternFill
from openpyxl.chart import (
    AreaChart,
    Reference,
    Series,
    LineChart,
    ScatterChart,
)

### Functions ###

'''
Function to create a CSV output file.
'''
def createCSV(fileName, dic):
    headers=['x', 'y']
    with open(fileName + '.csv', 'wb') as csvfile:
        writer = csv.writer(csvfile)
        writer.writerow(headers)
        for k, v in sorted(dic.items(), key=operator.itemgetter(0)):
            writer.writerow([k, v])

'''
Function to create a excel output file.
'''
def createEXCEL(fileName, dic):
    if fileName is None or dic is None:
        print("Error occurred")
        return

    headers=['x', 'y']
    wb = Workbook()
    ws = wb.active

    # Create table headers
    ws.append(headers)
        
    # Create ordered table
    for k, v in sorted(dic.items(), key=operator.itemgetter(0)):
            ws.append([k, v])

    # Center all cels
    for col in ws.columns:
        for cell in col:
            cell.alignment = Alignment(horizontal="center")

    # Border + background for headers
    thin_border = Border(left=Side(style='thin'), 
                     right=Side(style='thin'), 
                     top=Side(style='thin'), 
                     bottom=Side(style='thin'))
    for i in range(len(dic) + 1):
        for j in range(len(headers)):
            ws.cell(row=i+1, column=j+1).border = thin_border
            if i == 0:
                ws.cell(1, j+1).fill = PatternFill(start_color="FFC7CE", end_color="FFC7CE", fill_type = "solid")

    # Create graph
    chart = ScatterChart()
    chart.title = "LineChart"
    chart.style = 13
    chart.x_axis.title = 'X'
    chart.y_axis.title = 'Y'
    chart.legend = None
    x = Reference(ws, min_col=1, min_row=2, max_col=1, max_row=(len(dic)+1))
    y = Reference(ws, min_col=2, min_row=2, max_col=2, max_row=(len(dic)+1))
    s = Series(y, xvalues=x)
    chart.append(s)

    ws.add_chart(chart, "E4")
    wb.save(fileName + ".xlsx")